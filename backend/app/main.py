from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Form, Query, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
import os
import io
import uuid
import jwt
import shutil
import openpyxl
import pandas as pd
import zipfile
import re
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import quote

from . import models, schemas, crud
from .database import engine, get_db
from .pdf_parser import parse_project_pdf, fuzzy_match_project

# Ensure uploads directory exists
UPLOAD_DIR = Path(__file__).parent.parent / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

# In production use Alembic. Here we create tables directly on startup for simplicity.
models.Base.metadata.create_all(bind=engine)

# SQLite: add new columns idempotently (create_all does NOT alter existing tables)
def _migrate_add_columns():
    import sqlite3, os
    db_url = str(engine.url)
    if 'sqlite' not in db_url:
        return
    db_path = db_url.replace('sqlite:///', '').lstrip('/')
    if not os.path.exists(db_path):
        db_path = db_url.split('sqlite:///')[-1]
    if not os.path.exists(db_path):
        return
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("PRAGMA table_info(projects)")
    existing = {row[1] for row in cur.fetchall()}
    new_cols = [
        ("proposer",               "VARCHAR"),
        ("post_delivery_person",   "VARCHAR"),
        ("improvement_site",       "JSON"),
        ("owning_company",         "JSON"),
        ("improvement_purpose",    "JSON"),
        ("improvement_method",     "JSON"),
        ("needs_evaluation",       "VARCHAR"),
        ("implementation_period",  "INTEGER"),
        ("planned_start_date",     "DATETIME"),
        ("budget",                 "FLOAT"),
        ("budget_text",            "VARCHAR"),
        ("expected_revenue",       "FLOAT"),
        ("expected_revenue_text",  "VARCHAR"),
        ("is_one_time_investment", "BOOLEAN"),
        ("is_one_time_revenue",    "BOOLEAN"),
        ("quantitative_goal",      "TEXT"),
        ("current_problem",        "TEXT"),
        ("technical_solution",     "TEXT"),
        ("planned_end_date",       "DATETIME"),
    ]
    for col_name, col_type in new_cols:
        if col_name not in existing:
            cur.execute(f"ALTER TABLE projects ADD COLUMN {col_name} {col_type}")
            print(f"[migrate] Added column: projects.{col_name}")
    conn.commit()
    conn.close()

_migrate_add_columns()


app = FastAPI(title="技改项目智能调度中枢 API")

# CORS配置 - 开发环境允许所有来源
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,  # 使用通配符时不能为True
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve uploaded files statically
app.mount("/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")

JWT_SECRET = os.getenv("JWT_SECRET", "super-secret-key-123")

@app.get("/")
def root():
    return {"message": "Welcome to CompanyHub AI API"}

# ---- 项目主业务接口 ----

@app.post("/api/projects/import")
async def import_projects(file: UploadFile = File(...), db: Session = Depends(get_db)):
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(content), data_only=True)
        sheet = wb.active
        
        imported_count = 0
        skipped_count = 0

        # 表头行检测（支持第2行或第3行作为表头）
        header_row = None
        for row_idx in [2, 3]:  # 尝试第2行和第3行
            test_row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            if any('项目编号' in str(cell) or '项目名称' in str(cell) for cell in test_row if cell):
                header_row = row_idx
                break
        
        data_start_row = header_row + 1 if header_row else 4
        
        # 获取表头并建立列名映射
        headers = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0] if header_row else []
        header_map = {}
        for idx, h in enumerate(headers):
            if h:
                header_map[str(h).strip()] = idx
        
        # 查找关键字段的列索引（支持多种表头别名）
        def find_column(*aliases):
            for alias in aliases:
                if alias in header_map:
                    return header_map[alias]
            return None
        
        col_title = find_column('项目名称', '项目标题', 'name', 'title') or 4
        col_dept = find_column('申报单位', '部门', 'department', 'dept') or 1
        col_project_no = find_column('项目编号', '编号', 'project_no', 'no') or 0
        # 优先使用 F项目实施负责人/F项目主要参与人员，同时兼容之前的别名
        col_leader = find_column('F项目实施负责人', '项目实施负责人', '负责人', 'leader', '项目经理') or 39
        col_participants = find_column('F项目主要参与人员', '项目主要参与人员', '参与人员', 'participants', '成员') or 41
        
        # 标签列映射（从第7列开始）
        tag_columns = {
            7: "长材厂",
            8: "大棒厂", 
            9: "特钢厂",
            10: "涂层厂",
            11: "特冶厂",
            12: "电渣车间",
            13: "炼钢厂",
            14: "改造场地其他",
            16: "富钢",
            17: "盛特隆",
            18: "富佰",
            19: "隆毅",
            23: "工艺改善",
            24: "节能降耗",
            25: "质量提升",
            26: "成本优化",
            27: "产业提升",
            28: "安全环保",
            29: "布局优化",
            30: "信息化",
            34: "管理创新",
            35: "技术改造",
            36: "新产品新工艺",
            37: "战略性项目",
        }

        for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
            if not row:
                continue
            
            # 获取项目编号（D列）
            project_code = str(row[col_project_no]).strip() if col_project_no < len(row) and row[col_project_no] and str(row[col_project_no]).strip() not in ['None', ''] else None
            
            # 获取项目名称（关键字段）
            title = str(row[col_title]).strip() if row[col_title] and str(row[col_title]).strip() not in ['None', ''] else None
            if not title:
                continue
            
            # 获取部门
            department = str(row[col_dept]).strip() if row[col_dept] and str(row[col_dept]).strip() not in ['None', ''] else "未知部门"
            
            # 获取负责人
            leader = str(row[col_leader]).strip() if col_leader < len(row) and row[col_leader] and str(row[col_leader]).strip() not in ['None', ''] else None
            
            # 获取参与人员（解析逗号、顿号分隔的字符串）
            participants = []
            if col_participants < len(row) and row[col_participants]:
                participants_str = str(row[col_participants]).strip()
                if participants_str and participants_str != 'None':
                    # 支持顿号、逗号、空格分隔
                    import re
                    participants = [p.strip() for p in re.split('[、,，\s]+', participants_str) if p.strip()]
            
            # 构建标签列表
            tags = ["#实施中"]  # 默认标签
            
            for col_idx, tag_name in tag_columns.items():
                if col_idx < len(row) and row[col_idx]:
                    val = str(row[col_idx]).strip()
                    if val in ['勾选', '是', '✓', '√', 'True', '1']:
                        tags.append(f"#{tag_name}")
            
            # 检查是否已存在（按标题去重）
            existing = crud.get_project_by_title(db, title=title)
            if existing:
                skipped_count += 1
                continue
            
            # 创建项目
            project_data = schemas.ProjectCreate(
                title=title,
                project_code=project_code,
                department=department,
                leader=leader,
                participants=participants,
                tags=tags,
                status=models.ProjectStatus.in_progress
            )
            crud.create_project(db, project_data)
            imported_count += 1
            
        return {
            "message": f"解析完毕，成功导入 {imported_count} 个新项目，跳过 {skipped_count} 个已存在的项目。",
            "imported": imported_count,
            "skipped": skipped_count
        }
        
    except Exception as e:
        import traceback
        raise HTTPException(status_code=400, detail=f"文件解析失败: {str(e)}\n{traceback.format_exc()}")


@app.post("/api/projects/import-pdf")
async def import_projects_from_pdf(
    files: List[UploadFile] = File(...),
    db: Session = Depends(get_db)
):
    """
    批量导入 PDF 立项申请表（Q/FG G0286-2022 格式）。
    
    处理逻辑（每个 PDF）：
    1. 解析 PDF，提取所有字段（含勾选框）。
    2. 与现有项目名称做模糊匹配（≥80% 相似度）：
       - 匹配成功 → 将 PDF 文件存入该项目的"立项申请表"位置，并用 PDF 数据补全/更新项目字段。
       - 匹配失败 → 创建新项目，并将 PDF 文件作为立项申请表附件存入该项目。
    3. 返回处理摘要（成功/更新/新建/失败 各计数 + 明细）。
    """
    # 加载所有项目标题用于模糊匹配
    all_projects = crud.search_projects(db, keyword="", tags=[])
    title_map = {p.title: p for p in all_projects}   # 标题 → 项目对象

    results = []
    for upload_file in files:
        filename = upload_file.filename or "unknown.pdf"
        try:
            file_bytes = await upload_file.read()
            parsed = parse_project_pdf(file_bytes)
            pdf_title = parsed.get('title', '')

            # 模糊匹配
            matched_title = fuzzy_match_project(pdf_title, list(title_map.keys()), threshold=0.80)

            if matched_title:
                # ── 命中已有项目：按方案①合并更新 + 附加 PDF 文件 ──
                # 策略：
                #   1) 标量字段：PDF 解析出有值 → 覆盖老值；PDF 为 None/空 → 保留老值
                #   2) JSON 数组字段：合并去重（老值 ∪ PDF 值），保留顺序
                #   3) 日期字段：PDF 有值且能正确解析才覆盖
                existing = title_map[matched_title]
                update_data = {}

                scalar_fields = [
                    'proposer', 'post_delivery_person', 'needs_evaluation',
                    'implementation_period', 'leader', 'department',
                    'budget', 'budget_text',
                    'expected_revenue', 'expected_revenue_text',
                    'is_one_time_investment', 'is_one_time_revenue',
                    'quantitative_goal', 'current_problem', 'technical_solution',
                ]
                for f in scalar_fields:
                    v = parsed.get(f)
                    # 空串 / None 视为 PDF 未提供，不覆盖老值
                    if v is not None and not (isinstance(v, str) and v.strip() == ''):
                        update_data[f] = v

                # JSON 数组字段：合并去重（稳定顺序）
                def _merge_unique(old: list, new: list) -> list:
                    old = list(old or [])
                    seen = set(old)
                    for item in (new or []):
                        if item not in seen:
                            old.append(item)
                            seen.add(item)
                    return old

                for arr_f in ['improvement_site', 'owning_company',
                              'improvement_purpose', 'improvement_method',
                              'participants', 'tags']:
                    pdf_val = parsed.get(arr_f) or []
                    if pdf_val:
                        merged = _merge_unique(getattr(existing, arr_f, None) or [], pdf_val)
                        if merged != (getattr(existing, arr_f, None) or []):
                            update_data[arr_f] = merged

                # planned_start_date / planned_end_date：PDF 有值才覆盖
                for _date_field in ('planned_start_date', 'planned_end_date'):
                    if parsed.get(_date_field):
                        from datetime import datetime as _dt
                        try:
                            update_data[_date_field] = _dt.strptime(parsed[_date_field], '%Y-%m-%d')
                        except Exception:
                            pass

                # project_code：PDF 有值且老值为空时写入（避免把人工编号覆盖）
                if parsed.get('project_code') and not existing.project_code:
                    update_data['project_code'] = parsed['project_code']

                if update_data:
                    # 直接写库，绕过 PUT /api/projects/{id} 的字段级留痕
                    # （避免 PDF 批量导入刷屏，改由导入结束时聚合一条 pdf_import 日志）
                    for k, v in update_data.items():
                        setattr(existing, k, v)
                    if 'tags' in update_data:
                        crud.sync_status_from_tags(existing)
                    db.commit()
                    db.refresh(existing)

                # 存 PDF 文件
                _save_pdf_as_application(db, existing.id, file_bytes, filename)

                results.append({
                    "file": filename,
                    "action": "updated",
                    "project_id": str(existing.id),
                    "project_title": existing.title,
                    "matched_to": matched_title,
                })
            else:
                # ── 未命中：新建项目 ──
                from datetime import datetime as _dt
                planned_start = None
                planned_end = None
                if parsed.get('planned_start_date'):
                    try:
                        planned_start = _dt.strptime(parsed['planned_start_date'], '%Y-%m-%d')
                    except:
                        pass
                if parsed.get('planned_end_date'):
                    try:
                        planned_end = _dt.strptime(parsed['planned_end_date'], '%Y-%m-%d')
                    except:
                        pass

                new_project_data = schemas.ProjectCreate(
                    project_code=parsed.get('project_code'),
                    title=parsed.get('title', filename),
                    department=parsed.get('department'),
                    leader=parsed.get('leader'),
                    participants=parsed.get('participants', []),
                    tags=parsed.get('tags', ['#实施中']),
                    proposer=parsed.get('proposer'),
                    post_delivery_person=parsed.get('post_delivery_person'),
                    improvement_site=parsed.get('improvement_site', []),
                    owning_company=parsed.get('owning_company', []),
                    improvement_purpose=parsed.get('improvement_purpose', []),
                    improvement_method=parsed.get('improvement_method', []),
                    needs_evaluation=parsed.get('needs_evaluation'),
                    implementation_period=parsed.get('implementation_period'),
                    planned_start_date=planned_start,
                    planned_end_date=planned_end,
                    budget=parsed.get('budget'),
                    budget_text=parsed.get('budget_text'),
                    expected_revenue=parsed.get('expected_revenue'),
                    expected_revenue_text=parsed.get('expected_revenue_text'),
                    is_one_time_investment=parsed.get('is_one_time_investment'),
                    is_one_time_revenue=parsed.get('is_one_time_revenue'),
                    quantitative_goal=parsed.get('quantitative_goal'),
                    current_problem=parsed.get('current_problem'),
                    technical_solution=parsed.get('technical_solution'),
                )
                new_project = crud.create_project(db, new_project_data)

                # 存 PDF 文件
                _save_pdf_as_application(db, new_project.id, file_bytes, filename)

                # 加入 title_map 防止同批次重复创建
                title_map[new_project.title] = new_project

                results.append({
                    "file": filename,
                    "action": "created",
                    "project_id": str(new_project.id),
                    "project_title": new_project.title,
                })
        except Exception as e:
            import traceback
            results.append({
                "file": filename,
                "action": "error",
                "error": f"{str(e)}\n{traceback.format_exc()[:300]}",
            })

    created = sum(1 for r in results if r["action"] == "created")
    updated = sum(1 for r in results if r["action"] == "updated")
    errors  = sum(1 for r in results if r["action"] == "error")

    # ── 聚合留痕 ────────────────────────────────────────────
    # 按"这次导入了多少 PDF"汇总成一条干预动作，但挂在每一个被写入/新建的项目下，
    # 这样详情页看自己项目的时间轴能看到"本次导入立项表"，调度台全站流也能展开。
    summary_text = f"PDF 立项表导入：本次处理 {len(files)} 份，新建 {created}，更新 {updated}，失败 {errors}"
    touched_project_ids = {
        r.get("project_id") for r in results
        if r.get("action") in ("created", "updated") and r.get("project_id")
    }
    for pid_str in touched_project_ids:
        try:
            crud.log_change(
                db,
                uuid.UUID(pid_str),
                action_type='pdf_import',
                summary=summary_text,
                details={
                    "total": len(files),
                    "created": created,
                    "updated": updated,
                    "errors": errors,
                }
            )
        except Exception as _e:
            print(f"[pdf_import log] failed for {pid_str}: {_e}")

    return {
        "message": f"处理 {len(files)} 个 PDF：新建 {created} 个，更新 {updated} 个，失败 {errors} 个。",
        "created": created, "updated": updated, "errors": errors,
        "details": results,
    }


def _save_pdf_as_application(db, project_id, file_bytes: bytes, original_filename: str):
    """将 PDF 文件保存到磁盘并记录到 project_files 表（file_type=application）。"""
    import shutil
    from . import models as _models
    from .models import FileType

    # 如果已存在同名立项申请表则跳过
    existing_files = crud.get_project_files(db, project_id)
    for ef in existing_files:
        if getattr(ef.file_type, 'value', str(ef.file_type)) == 'application':
            return  # 已存在，不重复写入

    # 保存文件
    project_upload_dir = UPLOAD_DIR / str(project_id) / "application"
    project_upload_dir.mkdir(parents=True, exist_ok=True)
    safe_name = re.sub(r'[^\w\u4e00-\u9fff\-_.]', '_', original_filename)
    dest = project_upload_dir / safe_name
    dest.write_bytes(file_bytes)

    storage_path = str(dest.relative_to(UPLOAD_DIR.parent)).replace('\\', '/')
    file_record = _models.ProjectFile(
        project_id=project_id,
        file_type=FileType.application,
        storage_path=storage_path,
        original_name=original_filename,
        file_size=len(file_bytes),
        uploaded_by="system",
    )
    db.add(file_record)
    db.commit()


@app.post("/api/projects", response_model=schemas.ProjectResponse)

def create_project(project: schemas.ProjectCreate, db: Session = Depends(get_db)):
    return crud.create_project(db, project)

@app.get("/api/projects", response_model=List[schemas.ProjectResponse])
def read_projects(keyword: str = "", tags: str = "", db: Session = Depends(get_db)):
    tags_list = tags.split(",") if tags else []
    return crud.search_projects(db, keyword=keyword, tags=tags_list)

@app.get("/api/projects/{id}", response_model=schemas.ProjectResponse)
def read_project(id: uuid.UUID, db: Session = Depends(get_db)):
    project = crud.get_project(db, id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return project

_FIELD_LABELS = {
    'title':                '项目名称',
    'project_code':         '项目编号',
    'department':           '申报单位',
    'leader':               '项目负责人',
    'participants':         '参与人员',
    'proposer':             '项目提出者',
    'post_delivery_person': '交付后负责人',
    'planned_start_date':   '项目开始时间',
    'planned_end_date':     '项目结束时间',
    'end_date':             '项目结项时间',
    'current_problem':      '现状问题',
    'technical_solution':   '采取的措施',
    'quantitative_goal':    '实施效益分析',
    'needs_evaluation':     '需评审类型',
    'implementation_period': '实施周期',
    'budget':               '投入',
    'budget_text':          '投入',
    'expected_revenue':     '预计收益',
    'expected_revenue_text': '预计收益',
    'improvement_site':     '改造场地',
    'owning_company':       '所属公司',
    'improvement_purpose':  '改善目的',
    'improvement_method':   '改善方法',
}

def _fmt_val(v):
    if v is None or v == '':
        return '空'
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, (list, tuple)):
        return '、'.join(str(x) for x in v) if v else '空'
    return str(v)

def _values_equal(a, b):
    """宽松比较：None / [] / '' 视为相同空态；日期按日比较。"""
    def _empty(x):
        return x is None or x == '' or (isinstance(x, (list, tuple)) and len(x) == 0)
    if _empty(a) and _empty(b):
        return True
    if isinstance(a, datetime) and isinstance(b, datetime):
        return a.date() == b.date()
    if isinstance(a, (list, tuple)) and isinstance(b, (list, tuple)):
        return list(a) == list(b)
    return a == b


@app.put("/api/projects/{id}", response_model=schemas.ProjectResponse)
def update_project(id: uuid.UUID, updates: schemas.ProjectUpdate, db: Session = Depends(get_db)):
    db_project = crud.get_project(db, id)
    if not db_project:
        raise HTTPException(status_code=404, detail="Project not found")

    update_payload = updates.model_dump(exclude_none=True)

    # ── 结项时间（end_date）变更：强制要求延期原因，同时写 delay_history ──
    old_end_date = db_project.end_date
    new_end_date = update_payload.get('end_date')
    if new_end_date is not None and not _values_equal(old_end_date, new_end_date):
        reason = (update_payload.get('delay_reason') or '').strip()
        if not reason:
            raise HTTPException(status_code=400, detail="修改结项时间必须提供延期原因")
        crud.create_delay_history(
            db,
            project_id=id,
            old_end_date=old_end_date,
            new_end_date=new_end_date,
            reason=reason,
            changed_by="系统"
        )
        crud.log_change(
            db, id,
            action_type='date_delay',
            field_name='end_date',
            old_value=old_end_date, new_value=new_end_date,
            summary=f"结项时间 {_fmt_val(old_end_date)} → {_fmt_val(new_end_date)}（原因：{reason}）",
        )

    # ── planned_end_date 变更：同样需要原因，等价于延期 ──
    old_planned_end = db_project.planned_end_date
    new_planned_end = update_payload.get('planned_end_date')
    planned_end_changed = (
        new_planned_end is not None and not _values_equal(old_planned_end, new_planned_end)
    )
    if planned_end_changed:
        reason = (update_payload.get('delay_reason') or '').strip()
        if not reason:
            raise HTTPException(status_code=400, detail="修改项目结束时间必须提供变更原因")
        # 同样写入延期历史，方便复用前端 UI
        crud.create_delay_history(
            db,
            project_id=id,
            old_end_date=old_planned_end,
            new_end_date=new_planned_end,
            reason=reason,
            changed_by="系统"
        )
        crud.log_change(
            db, id,
            action_type='date_delay',
            field_name='planned_end_date',
            old_value=old_planned_end, new_value=new_planned_end,
            summary=f"项目结束时间 {_fmt_val(old_planned_end)} → {_fmt_val(new_planned_end)}（原因：{reason}）",
        )

    # ── 开始时间变更：自动留痕，不需要原因 ──
    old_start = db_project.planned_start_date
    new_start = update_payload.get('planned_start_date')
    if new_start is not None and not _values_equal(old_start, new_start):
        crud.log_change(
            db, id,
            action_type='date_edit',
            field_name='planned_start_date',
            old_value=old_start, new_value=new_start,
            summary=f"项目开始时间 {_fmt_val(old_start)} → {_fmt_val(new_start)}",
        )

    # ── 标签增删：对比旧/新 tags 列表 ──
    if 'tags' in update_payload:
        old_tags = list(db_project.tags or [])
        new_tags = list(update_payload.get('tags') or [])
        added = [t for t in new_tags if t not in old_tags]
        removed = [t for t in old_tags if t not in new_tags]
        for t in added:
            crud.log_change(db, id, action_type='tag_add', field_name='tags',
                            new_value=t, summary=f"新增标签 {t}")
        for t in removed:
            crud.log_change(db, id, action_type='tag_remove', field_name='tags',
                            old_value=t, summary=f"删除标签 {t}")

    # ── 状态字段变更 ──
    if 'status' in update_payload:
        old_status = getattr(db_project.status, 'value', str(db_project.status) if db_project.status else '')
        new_status_val = update_payload['status']
        new_status = getattr(new_status_val, 'value', str(new_status_val))
        if old_status != new_status:
            crud.log_change(db, id, action_type='status_change', field_name='status',
                            old_value=old_status, new_value=new_status,
                            summary=f"项目状态 {old_status or '空'} → {new_status}")

    # ── 普通字段编辑（标量 / 数组）────────────────────────────
    tracked_fields = [
        'title', 'project_code', 'department', 'leader', 'participants',
        'proposer', 'post_delivery_person',
        'current_problem', 'technical_solution', 'quantitative_goal',
        'needs_evaluation', 'implementation_period',
        'budget', 'budget_text', 'expected_revenue', 'expected_revenue_text',
        'improvement_site', 'owning_company', 'improvement_purpose', 'improvement_method',
    ]
    for f in tracked_fields:
        if f not in update_payload:
            continue
        old_v = getattr(db_project, f, None)
        new_v = update_payload[f]
        if _values_equal(old_v, new_v):
            continue
        label = _FIELD_LABELS.get(f, f)
        crud.log_change(
            db, id,
            action_type='field_edit',
            field_name=f,
            old_value=old_v, new_value=new_v,
            summary=f"{label}：{_fmt_val(old_v)} → {_fmt_val(new_v)}",
        )

    project = crud.update_project(db, id, updates)
    return project


@app.get("/api/projects/{id}/delay-history", response_model=List[schemas.ProjectDelayHistoryResponse])
def get_project_delay_history(id: uuid.UUID, db: Session = Depends(get_db)):
    """获取项目延期历史"""
    project = crud.get_project(db, id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return crud.get_delay_history(db, id)

_STATE_TAGS = {'#暂停', '#暂停中', '#已完成', '#已结项', '#实施中', '#待结项'}

def _diagnose_status_change(old_status: str, new_status: str) -> Optional[str]:
    """把"状态枚举变了"翻译成用户看得懂的一句话。"""
    if old_status == new_status:
        return None
    if new_status == '暂停中':
        return "状态：强制暂停"
    if old_status == '暂停中' and new_status == '实施中':
        return "状态：恢复实施"
    if new_status == '已完成':
        return "状态：标记为已完成（待结项）"
    if new_status == '已结项':
        return "状态：归档结项"
    return f"状态：{old_status or '空'} → {new_status}"

@app.put("/api/projects/{id}/complete", response_model=schemas.ProjectResponse)
def complete_project(id: uuid.UUID, db: Session = Depends(get_db)):
    before = crud.get_project(db, id)
    if not before:
        raise HTTPException(status_code=404, detail="Project not found")
    old_status = before.status.value if before.status else ''
    project = crud.complete_project(db, id)
    new_status = project.status.value if project.status else ''
    if old_status != new_status:
        crud.log_change(
            db, id,
            action_type='status_change',
            field_name='status',
            old_value=old_status, new_value=new_status,
            summary=_diagnose_status_change(old_status, new_status) or f"状态：{old_status} → {new_status}",
        )
    return project

@app.post("/api/projects/{id}/tags", response_model=schemas.ProjectResponse)
def add_project_tag(id: uuid.UUID, payload: dict, db: Session = Depends(get_db)):
    """添加标签到项目"""
    project = crud.get_project(db, id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    tag = payload.get("tag", "").strip()
    if not tag:
        raise HTTPException(status_code=400, detail="标签不能为空")

    if not tag.startswith("#"):
        tag = "#" + tag

    current_tags = list(project.tags or [])
    if tag in current_tags:
        raise HTTPException(status_code=400, detail="标签已存在")

    old_status = project.status.value if project.status else ''
    current_tags.append(tag)
    project.tags = current_tags
    crud.sync_status_from_tags(project)
    db.commit()
    db.refresh(project)
    new_status = project.status.value if project.status else ''

    # 只有非状态标签才算"新增标签"；状态标签用 status_change 统一留痕
    if tag in _STATE_TAGS:
        if old_status != new_status:
            crud.log_change(
                db, id,
                action_type='status_change',
                field_name='status',
                old_value=old_status, new_value=new_status,
                summary=_diagnose_status_change(old_status, new_status) or f"状态：{old_status} → {new_status}",
            )
    else:
        crud.log_change(db, id, action_type='tag_add', field_name='tags',
                        new_value=tag, summary=f"新增标签 {tag}")

    return project

@app.delete("/api/projects/{id}/tags/{tag}", response_model=schemas.ProjectResponse)
def remove_project_tag(id: uuid.UUID, tag: str, db: Session = Depends(get_db)):
    """从项目中删除标签"""
    project = crud.get_project(db, id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    from urllib.parse import unquote
    tag = unquote(tag)

    current_tags = list(project.tags or [])
    if tag not in current_tags:
        raise HTTPException(status_code=404, detail="标签不存在")

    old_status = project.status.value if project.status else ''
    current_tags.remove(tag)
    project.tags = current_tags
    crud.sync_status_from_tags(project)
    db.commit()
    db.refresh(project)
    new_status = project.status.value if project.status else ''

    if tag in _STATE_TAGS:
        if old_status != new_status:
            crud.log_change(
                db, id,
                action_type='status_change',
                field_name='status',
                old_value=old_status, new_value=new_status,
                summary=_diagnose_status_change(old_status, new_status) or f"状态：{old_status} → {new_status}",
            )
    else:
        crud.log_change(db, id, action_type='tag_remove', field_name='tags',
                        old_value=tag, summary=f"删除标签 {tag}")

    return project

# ---- 文件管理接口 ----

@app.get("/api/projects/{id}/files", response_model=List[schemas.ProjectFileResponse])
def get_project_files(id: uuid.UUID, db: Session = Depends(get_db)):
    """获取项目下的所有文件列表"""
    project = crud.get_project(db, id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return crud.get_project_files(db, id)

@app.post("/api/projects/{id}/files", response_model=schemas.ProjectFileResponse)
async def upload_project_file(
    id: uuid.UUID,
    file_type: str = Form(...),
    uploaded_by: str = Form("匿名"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    上传文件到项目资料池
    存储路径规范：uploads/{project_id}/{file_type}/{filename}
    """
    project = crud.get_project(db, id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # 验证并转换 file_type
    type_map = {
        "application": models.FileType.application,
        "ppt": models.FileType.ppt,
        "free_resource": models.FileType.free_resource,
    }
    ft = type_map.get(file_type)
    if not ft:
        raise HTTPException(status_code=400, detail=f"不支持的文件类型: {file_type}，可选: application, ppt, free_resource")

    # 获取原始文件名和扩展名
    original_name = file.filename or "unknown"
    ext = Path(original_name).suffix.lower()
    safe_filename = f"{uuid.uuid4().hex}{ext}"

    # 构建分层存储路径：uploads/{project_id}/{file_type}/
    project_dir = UPLOAD_DIR / str(id) / file_type
    project_dir.mkdir(parents=True, exist_ok=True)
    
    storage_path = project_dir / safe_filename
    relative_path = f"uploads/{id}/{file_type}/{safe_filename}"

    # 读取文件内容并保存
    content = await file.read()
    file_size = len(content)
    
    with open(storage_path, "wb") as buffer:
        buffer.write(content)

    # 保存到数据库
    file_data = schemas.ProjectFileCreate(
        file_type=ft,
        original_name=original_name
    )
    db_file = crud.add_project_file_v2(db, id, file_data, relative_path, file_size, uploaded_by)

    # 文件上传留痕（每个文件一条）
    _file_type_label = {
        'application': '立项申请表',
        'ppt': '结项PPT',
        'free_resource': '自由资料',
    }.get(file_type, file_type)
    crud.log_change(
        db, id,
        action_type='file_upload',
        field_name=file_type,
        new_value=original_name,
        summary=f"上传{_file_type_label}：{original_name}",
    )

    # 如果是结项PPT，自动更新项目状态
    if ft == models.FileType.ppt:
        old_status = project.status.value if project.status else ''
        tags = list(project.tags or [])
        if "#已完成" not in tags:
            tags.append("#已完成")
        project.tags = tags
        crud.sync_status_from_tags(project)
        db.commit()
        new_status = project.status.value if project.status else ''
        if old_status != new_status:
            crud.log_change(
                db, id,
                action_type='status_change',
                field_name='status',
                old_value=old_status, new_value=new_status,
                summary=_diagnose_status_change(old_status, new_status) or f"状态：{old_status} → {new_status}",
            )

    return db_file

@app.get("/api/files/{file_id}/download")
def download_single_file(file_id: uuid.UUID, db: Session = Depends(get_db)):
    """单文件下载"""
    file_record = crud.get_file_by_id(db, file_id)
    if not file_record:
        raise HTTPException(status_code=404, detail="文件不存在")
    
    file_path = UPLOAD_DIR.parent / file_record.storage_path
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="文件已丢失")
    
    # RFC 5987 编码中文文件名
    encoded_filename = quote(file_record.original_name, safe='')
    
    return StreamingResponse(
        open(file_path, "rb"),
        media_type="application/octet-stream",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
            "Content-Length": str(file_record.file_size)
        }
    )

@app.post("/api/files/download/batch")
def download_batch_files(file_ids: List[str], db: Session = Depends(get_db)):
    """
    批量下载文件（打包成ZIP）
    file_ids: 文件ID列表
    """
    if not file_ids or len(file_ids) == 0:
        raise HTTPException(status_code=400, detail="请选择要下载的文件")
    
    # 获取所有文件记录
    files_to_zip = []
    for fid in file_ids:
        try:
            file_record = crud.get_file_by_id(db, uuid.UUID(fid))
            if file_record:
                file_path = UPLOAD_DIR.parent / file_record.storage_path
                if file_path.exists():
                    files_to_zip.append({
                        "path": file_path,
                        "name": file_record.original_name
                    })
        except:
            continue
    
    if len(files_to_zip) == 0:
        raise HTTPException(status_code=404, detail="未找到可下载的文件")
    
    # 单文件直接返回
    if len(files_to_zip) == 1:
        file_info = files_to_zip[0]
        encoded_name = quote(file_info["name"], safe='')
        return StreamingResponse(
            open(file_info["path"], "rb"),
            media_type="application/octet-stream",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}"}
        )
    
    # 多文件打包成 ZIP（一次性构建后返回）
    # 说明：旧实现的伪流式会破坏 ZIP 中央目录偏移，
    # 在宽容的 Windows 解压器里能用，但在浏览器/Linux unzip 严格校验时会报错。
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED, allowZip64=True) as zf:
        for file_info in files_to_zip:
            zf.write(file_info["path"], file_info["name"])
    buffer.seek(0)
    zip_size = buffer.getbuffer().nbytes

    encoded_filename = quote("下载文件.zip", safe='')

    return StreamingResponse(
        buffer,
        media_type="application/zip",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
            "Content-Length": str(zip_size),
            "X-Accel-Buffering": "no",
        },
    )

@app.delete("/api/files/{file_id}")
def delete_file(file_id: uuid.UUID, db: Session = Depends(get_db)):
    """删除文件"""
    file_record = crud.get_file_by_id(db, file_id)
    if not file_record:
        raise HTTPException(status_code=404, detail="文件不存在")

    pid = file_record.project_id
    file_type_val = getattr(file_record.file_type, 'value', str(file_record.file_type))
    original_name = file_record.original_name

    # 删除物理文件
    file_path = UPLOAD_DIR.parent / file_record.storage_path
    if file_path.exists():
        file_path.unlink()

    # 删除数据库记录
    success = crud.delete_file(db, file_id)
    if not success:
        raise HTTPException(status_code=500, detail="删除失败")

    _file_type_label = {
        'application': '立项申请表',
        'ppt': '结项PPT',
        'free_resource': '自由资料',
    }.get(file_type_val, file_type_val)
    crud.log_change(
        db, pid,
        action_type='file_delete',
        field_name=file_type_val,
        old_value=original_name,
        summary=f"删除{_file_type_label}：{original_name}",
    )

    return {"message": "删除成功"}

# ---- 周期更新记录接口（原里程碑接口） ----

@app.get("/api/projects/{id}/milestones", response_model=List[schemas.MilestoneResponse])
def get_project_milestones(id: uuid.UUID, db: Session = Depends(get_db)):
    """获取项目的周期更新记录（里程碑）"""
    project = crud.get_project(db, id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return crud.get_milestones(db, id)

@app.post("/api/projects/{id}/milestones", response_model=schemas.MilestoneResponse)
def create_project_milestone(id: uuid.UUID, update: schemas.MilestoneCreate, db: Session = Depends(get_db)):
    """创建周期更新记录（里程碑）"""
    project = crud.get_project(db, id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return crud.create_milestone(db, id, update)

@app.delete("/api/milestones/{milestone_id}")
def delete_project_milestone(milestone_id: uuid.UUID, db: Session = Depends(get_db)):
    """删除周期更新记录（里程碑）"""
    success = crud.delete_milestone(db, milestone_id)
    if not success:
        raise HTTPException(status_code=404, detail="Milestone not found")
    return {"message": "删除成功"}

# ---- 二维码与权限沙盒 ----

@app.post("/api/projects/{id}/generate_qr")
def generate_qr(id: uuid.UUID, db: Session = Depends(get_db)):
    project = crud.get_project(db, id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    expire = datetime.utcnow() + timedelta(days=7)
    token = jwt.encode({"sub": str(id), "exp": expire}, JWT_SECRET, algorithm="HS256")
    url = f"/guest/project/{token}"
    return {"qr_url": url, "token": token}

@app.get("/api/guest/projects/{token}", response_model=schemas.ProjectResponse)
def guest_read_project(token: str, db: Session = Depends(get_db)):
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
        project_id = uuid.UUID(payload.get("sub"))
        return read_project(project_id, db)
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ---- AI 审批接口 ----

@app.get("/api/ai_actions/pending", response_model=List[schemas.AIPendingActionResponse])
def get_pending_actions(db: Session = Depends(get_db)):
    return crud.get_pending_actions(db)

@app.post("/api/ai_actions/{id}/approve", response_model=schemas.AIPendingActionResponse)
def approve_action(id: uuid.UUID, payload: dict, db: Session = Depends(get_db)):
    action = crud.approve_action(db, id, payload)
    if not action:
        raise HTTPException(status_code=404, detail="Action not found")
    
    if action.action_type == models.ActionType.add_milestone:
        deadline_str = payload.get("deadline")
        deadline = datetime.fromisoformat(deadline_str) if deadline_str else None
        
        milestone_data = schemas.MilestoneCreate(
            title=payload.get("title", "未命名里程碑"),
            deadline=deadline
        )
        crud.create_milestone(db, action.project_id, milestone_data)
    
    return action

@app.delete("/api/ai_actions/{id}")
def delete_action(id: uuid.UUID, db: Session = Depends(get_db)):
    success = crud.delete_action(db, id)
    if not success:
        raise HTTPException(status_code=404, detail="Action not found")
    return {"message": "Deleted"}

# ---- 项目固化记录接口（Timeline Snapshot） ----

@app.post("/api/upload/images")
async def upload_images(files: List[UploadFile] = File(...)):
    """
    批量上传图片接口
    存储路径：uploads/images/{uuid}.{ext}
    返回：图片URL列表
    """
    # 确保图片目录存在
    images_dir = UPLOAD_DIR / "images"
    images_dir.mkdir(parents=True, exist_ok=True)
    
    uploaded_urls = []
    
    for file in files:
        if not file.content_type or not file.content_type.startswith("image/"):
            raise HTTPException(status_code=400, detail=f"文件 {file.filename} 不是图片格式")
        
        # 生成唯一文件名
        ext = Path(file.filename or "").suffix.lower() or ".jpg"
        if ext not in ['.jpg', '.jpeg', '.png', '.gif', '.webp']:
            ext = '.jpg'
        safe_filename = f"{uuid.uuid4().hex}{ext}"
        
        # 保存文件
        file_path = images_dir / safe_filename
        content = await file.read()
        
        with open(file_path, "wb") as buffer:
            buffer.write(content)
        
        # 返回相对URL
        uploaded_urls.append(f"/uploads/images/{safe_filename}")
    
    return {"image_urls": uploaded_urls}

@app.get("/api/projects/{id}/updates", response_model=List[schemas.ProjectUpdateResponse])
def get_project_updates(id: uuid.UUID, db: Session = Depends(get_db)):
    """获取项目的固化记录列表（按时间倒序）"""
    project = crud.get_project(db, id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return crud.get_project_updates(db, id)

@app.get("/api/history", response_model=List[schemas.HistoryItemResponse])
def get_global_history(limit: int = Query(50, description="限制返回的数量"), db: Session = Depends(get_db)):
    """获取全站最新项目固化纪录"""
    rows = crud.get_all_history(db, limit=limit)
    res = []
    for update_obj, project_title in rows:
        item_dict = {
            "id": update_obj.id,
            "project_id": update_obj.project_id,
            "reporter_name": update_obj.reporter_name,
            "content": update_obj.content,
            "image_urls": update_obj.image_urls,
            "remarks": update_obj.remarks,
            "created_at": update_obj.created_at,
            "project_title": project_title
        }
        res.append(item_dict)
    return res

@app.post("/api/history/{id}/remarks", response_model=schemas.ProjectLogRemarkResponse)
def create_history_remark(id: uuid.UUID, payload: schemas.HistoryRemarkCreate, db: Session = Depends(get_db)):
    """向某个历史固化记录追加备注"""
    # 验证记录是否存在
    update_record = crud.get_project_update(db, id)
    if not update_record:
        raise HTTPException(status_code=404, detail="History update record not found")
    
    new_remark = crud.create_history_remark(db, update_id=id, content=payload.content)
    return new_remark

@app.post("/api/projects/{id}/updates", response_model=schemas.ProjectUpdateResponse, status_code=201)
def create_project_update(id: uuid.UUID, update: schemas.ProjectUpdateCreate, db: Session = Depends(get_db)):
    """创建项目固化记录"""
    project = crud.get_project(db, id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # 验证：必须有内容或图片
    if not update.content.strip() and not update.image_urls:
        raise HTTPException(status_code=400, detail="汇报内容和图片不能同时为空")
    
    return crud.create_project_update(db, id, update)

# ---- AI 配置中心接口 (通义千问) ----

@app.get("/api/ai-configs", response_model=List[schemas.AIConfigListResponse])
def get_ai_configs(db: Session = Depends(get_db)):
    """获取AI配置列表（隐藏API Key）"""
    return crud.get_ai_configs(db)

@app.get("/api/ai-configs/{config_id}", response_model=schemas.AIConfigResponse)
def get_ai_config(config_id: uuid.UUID, db: Session = Depends(get_db)):
    """获取单个AI配置详情"""
    config = crud.get_ai_config(db, config_id)
    if not config:
        raise HTTPException(status_code=404, detail="Config not found")
    return config

@app.post("/api/ai-configs", response_model=schemas.AIConfigResponse)
def create_ai_config(config: schemas.AIConfigCreate, db: Session = Depends(get_db)):
    """创建AI配置"""
    return crud.create_ai_config(db, config)

@app.put("/api/ai-configs/{config_id}", response_model=schemas.AIConfigResponse)
def update_ai_config(config_id: uuid.UUID, updates: schemas.AIConfigUpdate, db: Session = Depends(get_db)):
    """更新AI配置"""
    config = crud.update_ai_config(db, config_id, updates)
    if not config:
        raise HTTPException(status_code=404, detail="Config not found")
    return config

@app.delete("/api/ai-configs/{config_id}")
def delete_ai_config(config_id: uuid.UUID, db: Session = Depends(get_db)):
    """删除AI配置"""
    success = crud.delete_ai_config(db, config_id)
    if not success:
        raise HTTPException(status_code=404, detail="Config not found")
    return {"message": "删除成功"}

@app.post("/api/ai-configs/{config_id}/activate")
def activate_ai_config(config_id: uuid.UUID, db: Session = Depends(get_db)):
    """激活指定AI配置"""
    config = crud.set_active_config(db, config_id)
    if not config:
        raise HTTPException(status_code=404, detail="Config not found")
    return config

@app.get("/api/ai-configs/active", response_model=schemas.AIConfigResponse)
def get_active_ai_config(db: Session = Depends(get_db)):
    """获取当前启用的AI配置"""
    config = crud.get_active_ai_config(db)
    if not config:
        raise HTTPException(status_code=404, detail="No active config found")
    return config

# ---- 数据导出接口 ----

# 标签分组（用于旧数据反推字段，兼容从表二导入的历史项目）
_SITE_TAGS    = {'#长材厂','#大棒厂','#特钢厂','#涂层厂','#特冶厂','#电渣车间','#炼钢厂','#改造场地其他'}
_COMPANY_TAGS = {'#富钢','#盛特隆','#富佰','#隆毅'}
_PURPOSE_TAGS = {'#工艺改善','#节能降耗','#质量提升','#成本优化','#产业提升','#安全环保','#布局优化','#信息化'}
_METHOD_TAGS  = {'#管理创新','#技术改造','#新产品新工艺','#战略性项目'}

def _tags_to(tags_list, tag_set):
    return [t.lstrip('#') for t in (tags_list or []) if t in tag_set]

def _project_has_ppt(db, project_id):
    files = crud.get_project_files(db, project_id)
    return any(getattr(f.file_type, 'value', str(f.file_type)) == 'ppt' for f in files)

@app.get("/api/projects/export/excel")
def export_projects_excel(
    keyword: str = "",
    tags: str = "",
    db: Session = Depends(get_db)
):
    """按表一（2026年度管理创新和技术改造项目明细汇总表）格式导出 Excel，共32列。"""
    try:
        tags_list = [t for t in tags.split(",") if t] if tags else []
        projects = crud.search_projects(db, keyword=keyword, tags=tags_list)
        # ProjectStatus.value 本身就是中文（见 models.ProjectStatus）
        # 旧实现用英文 key 查中文 value，永远查不到，导致"完成情况"列一直空
        # 这里直接用 value（已中文），同时兜底一层英文 key 以兼容脏数据
        STATUS_MAP = {
            '实施中': '实施中', '暂停中': '暂停中',
            '已完成': '已完成', '已结项': '已结项',
            'in_progress': '实施中', 'paused': '暂停中',
            'pending_completion': '已完成', 'completed': '已结项',
        }
        data = []
        for seq, p in enumerate(projects, start=1):
            tags_arr = list(p.tags or [])
            site_list   = list(p.improvement_site or []) or _tags_to(tags_arr, _SITE_TAGS)
            method_list = list(p.improvement_method or []) or _tags_to(tags_arr, _METHOD_TAGS)
            method_str  = method_list[0] if method_list else ''
            is_strategic = '是' if ('战略性项目' in method_list or '#战略性项目' in tags_arr) else '否'
            purp_list   = list(p.improvement_purpose or []) or _tags_to(tags_arr, _PURPOSE_TAGS)
            status_str  = STATUS_MAP.get(p.status.value if p.status else '', '')
            has_ppt     = _project_has_ppt(db, p.id)
            data.append({
                '月':               p.created_at.year % 100 if p.created_at else '',
                '序号':             seq,
                '项目编号':         p.project_code or '',
                '申报单位':         p.department or '',
                '申报时间':         p.created_at.strftime('%Y-%m-%d') if p.created_at else '',
                '项目名称':         p.title,
                '改造场地':         '、'.join(site_list),
                '项目负责人':       p.leader or '',
                '项目成员':         '、'.join(p.participants or []),
                '项目交付后负责人': p.post_delivery_person or '',
                '项目提出者':       p.proposer or '',
                '项目开始时间':     p.planned_start_date.strftime('%Y-%m-%d') if p.planned_start_date else '',
                # 项目结束时间优先使用 planned_end_date（新字段），兼容老数据回退到 end_date
                '项目结束时间':     (p.planned_end_date or p.end_date).strftime('%Y-%m-%d') if (p.planned_end_date or p.end_date) else '',
                '项目类别改善方法': method_str,
                '是否为战略项目':   is_strategic,
                '项目类别改善目的': purp_list[0] if len(purp_list) > 0 else '',
                '项目类别改善目的2': purp_list[1] if len(purp_list) > 1 else '',
                # 优先使用 PDF 原始文本（能保留 ">1"、"约 5" 这类表达），
                # 其次回退到纯数字字段，再没有就留空
                '投入（万元）':     (p.budget_text or (p.budget if p.budget is not None else '')),
                '开源收益（万元/年）': '',
                '节流收益（万元/年）': (p.expected_revenue_text or (p.expected_revenue if p.expected_revenue is not None else '')),
                '一次性投入':       '是' if p.is_one_time_investment else ('否' if p.is_one_time_investment is False else ''),
                '一次性收益':       '是' if p.is_one_time_revenue   else ('否' if p.is_one_time_revenue   is False else ''),
                '实施效益分析':     p.quantitative_goal or '',
                '现状问题':         p.current_problem or '',
                '采取的措施':       p.technical_solution or '',
                '完成情况':         status_str,
                '验收表进度':       '已收到' if has_ppt else '',
                '是否完成验收':     '是' if '#已结项' in tags_arr else '否',
                '可否形成专利':     '',
                '专利进度':         '',
                '备注':             '',
                'Q':                '',
            })
        df = pd.DataFrame(data)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='2026年度技改创新汇总表', index=False)
            ws = writer.sheets['2026年度技改创新汇总表']
            col_widths = [6,6,16,14,12,36,16,14,22,12,12,12,12,12,8,12,12,10,10,10,8,8,40,40,40,10,10,8,8,8,6,6]
            from openpyxl.utils import get_column_letter
            for i, w in enumerate(col_widths):
                ws.column_dimensions[get_column_letter(i+1)].width = w
            ws.freeze_panes = 'A2'
        output.seek(0)
        filename = f"项目汇总表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        encoded_filename = quote(filename, safe='')
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )
    except Exception as e:
        import traceback
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}\n{traceback.format_exc()}")




@app.get("/api/projects/export/zip")
def export_projects_zip(
    project_id: Optional[str] = Query(None, description="指定项目ID，不指定则导出所有"),
    tag: Optional[str] = Query(None, description="按标签筛选"),
    db: Session = Depends(get_db)
):
    """
    打包下载项目文件（流式传输，防止内存溢出）
    按项目名称自动创建文件夹结构
    """
    try:
        # 获取项目列表
        if project_id:
            try:
                pid = uuid.UUID(project_id)
                project = crud.get_project(db, pid)
                projects = [project] if project else []
            except:
                projects = []
        else:
            projects = crud.search_projects(db, keyword="", tags=[tag] if tag else [])
        
        # 如果未找到项目，依然允许流转到下方，因为 files_to_zip 将为空，我们能优雅返回包含 README 的 ZIP
        # 收集所有需要打包的文件
        files_to_zip = []
        for project in projects:
            project_files = crud.get_project_files(db, project.id)
            for f in project_files:
                file_path = UPLOAD_DIR.parent / f.storage_path
                if file_path.exists():
                    # 安全的文件夹名称（移除非法字符）
                    safe_folder = re.sub(r'[<>:"/\\|?*]', '_', project.title)[:50]
                    files_to_zip.append({
                        "path": file_path,
                        "arcname": f"{safe_folder}/{f.file_type.value}/{f.original_name}"
                    })
        
        # ── 一次性构建 ZIP 后整体返回 ──────────────────────────
        # 旧实现伪流式：在 ZipFile 未 close 前就 buffer.truncate()，
        # 导致中央目录（central directory）偏移错乱。
        # Windows 资源管理器/7-Zip 宽容能自愈，但浏览器/Linux unzip 严格校验就报错，
        # 这正是"本地能解、云端不行"的根因。
        #
        # 几十~几百 MB 的项目资料一次性在内存里构建完全可控，
        # 直接用 BytesIO 构建完整 ZIP 再返回，保证格式正确。
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED, allowZip64=True) as zf:
            if not files_to_zip:
                zf.writestr(
                    "说明_README.txt",
                    "该选择范围空空如也，无任何资料可供下载（No files found for your selection）.",
                )
            else:
                for file_info in files_to_zip:
                    zf.write(file_info["path"], file_info["arcname"])
        buffer.seek(0)
        zip_size = buffer.getbuffer().nbytes

        if project_id and len(projects) == 1:
            safe_name = re.sub(r'[<>:"/\\|?*]', '_', projects[0].title)[:30]
            filename = f"{safe_name}_文件包.zip"
        else:
            filename = f"项目文件包_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"

        encoded_filename = quote(filename, safe='')

        return StreamingResponse(
            buffer,
            media_type="application/zip",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
                "Content-Length": str(zip_size),
                # 云端 nginx 开了 proxy_buffering 时，告诉它不要 gzip/修改 zip 响应
                "X-Accel-Buffering": "no",
            },
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"打包下载失败: {str(e)}")


@app.get("/api/projects/{id}/export/zip")
def export_single_project_zip(id: uuid.UUID, db: Session = Depends(get_db)):
    """
    导出单个项目的所有文件为ZIP
    """
    project = crud.get_project(db, id)
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    
    return export_projects_zip(project_id=str(id), db=db)


# ---- 外部协同门户 JWT 认证 ----

class PortalTokenPayload:
    """门户 JWT Token 数据结构"""
    def __init__(self, project_ids: List[str], exp: datetime, iat: datetime):
        self.project_ids = project_ids
        self.exp = exp
        self.iat = iat

def create_portal_token(project_ids: List[str], expires_days: int = 7) -> str:
    """
    创建外部协同门户 JWT Token
    """
    now = datetime.utcnow()
    payload = {
        "sub": "portal",  # 主题标识
        "project_ids": project_ids,  # 可访问的项目ID列表
        "iat": now,  # 签发时间
        "exp": now + timedelta(days=expires_days),  # 过期时间
        "type": "portal"  # Token 类型
    }
    return jwt.encode(payload, JWT_SECRET, algorithm="HS256")

def verify_portal_token(token: str) -> PortalTokenPayload:
    """
    验证门户 JWT Token
    """
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
        
        # 验证 Token 类型
        if payload.get("type") != "portal":
            raise HTTPException(status_code=401, detail="无效的 Token 类型")
        
        return PortalTokenPayload(
            project_ids=payload.get("project_ids", []),
            exp=datetime.fromtimestamp(payload["exp"]),
            iat=datetime.fromtimestamp(payload["iat"])
        )
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token 已过期")
    except jwt.JWTError:
        raise HTTPException(status_code=401, detail="无效的 Token")

async def get_portal_auth(authorization: Optional[str] = Header(None)) -> PortalTokenPayload:
    """
    FastAPI 依赖：从请求头获取并验证门户 Token
    """
    if not authorization:
        raise HTTPException(status_code=401, detail="缺少认证信息")
    
    # 支持 Bearer token 格式
    if authorization.startswith("Bearer "):
        token = authorization[7:]
    else:
        token = authorization
    
    return verify_portal_token(token)

@app.post("/api/portal/auth")
def portal_login(credentials: dict, db: Session = Depends(get_db)):
    """
    外部协同门户登录接口 - 极简模式：只输入姓名即可查看
    请求体: { "name": "姓名" }
    """
    name = credentials.get("name", "").strip()
    
    if not name:
        raise HTTPException(status_code=400, detail="姓名不能为空")
    
    # 查找该姓名关联的项目（通过负责人或参与人员匹配）
    all_projects = crud.search_projects(db, keyword="", tags=[])
    accessible_projects = []
    
    for project in all_projects:
        # 匹配姓名（模糊匹配）
        leader_match = project.leader and name in project.leader
        participant_match = any(name in str(p) for p in (project.participants or []))
        if leader_match or participant_match:
            accessible_projects.append(str(project.id))
    
    # 生成 JWT Token
    token = create_portal_token(accessible_projects, expires_days=7)
    
    return {
        "token": token,
        "expires_in": 7 * 24 * 3600,  # 7天，单位秒
        "project_ids": accessible_projects,
        "name": name
    }

@app.get("/api/portal/projects", response_model=List[schemas.ProjectResponse])
def get_portal_projects(auth: PortalTokenPayload = Depends(get_portal_auth), db: Session = Depends(get_db)):
    """
    获取当前门户用户可访问的项目列表
    """
    if not auth.project_ids:
        return []  # 没有可访问的项目
    
    projects = []
    for pid_str in auth.project_ids:
        try:
            project = crud.get_project(db, uuid.UUID(pid_str))
            if project:
                projects.append(project)
        except:
            continue
    
    return projects

@app.get("/api/portal/projects/{id}", response_model=schemas.ProjectResponse)
def get_portal_project_detail(id: uuid.UUID, auth: PortalTokenPayload = Depends(get_portal_auth), db: Session = Depends(get_db)):
    """
    获取单个项目详情（门户用户）
    """
    # 验证是否有权限访问该项目
    if str(id) not in auth.project_ids:
        raise HTTPException(status_code=403, detail="无权访问该项目")
    
    project = crud.get_project(db, id)
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    
    return project

@app.get("/api/portal/projects/{id}/files", response_model=List[schemas.ProjectFileResponse])
def get_portal_project_files(id: uuid.UUID, auth: PortalTokenPayload = Depends(get_portal_auth), db: Session = Depends(get_db)):
    """
    获取项目文件列表（门户用户）
    """
    if str(id) not in auth.project_ids:
        raise HTTPException(status_code=403, detail="无权访问该项目")
    
    return crud.get_project_files(db, id)

@app.get("/api/portal/projects/{id}/updates", response_model=List[schemas.ProjectUpdateResponse])
def get_portal_project_updates(id: uuid.UUID, auth: PortalTokenPayload = Depends(get_portal_auth), db: Session = Depends(get_db)):
    """
    获取项目固化记录（门户用户）
    """
    if str(id) not in auth.project_ids:
        raise HTTPException(status_code=403, detail="无权访问该项目")
    
    return crud.get_project_updates(db, id)

@app.patch("/api/portal/projects/{id}/delivery-person", response_model=schemas.ProjectResponse)
def portal_update_delivery_person(
    id: uuid.UUID,
    payload: schemas.PortalDeliveryPersonUpdate,
    auth: PortalTokenPayload = Depends(get_portal_auth),
    db: Session = Depends(get_db),
):
    """
    外部门户用户修改项目"交付后负责人"。
    这是目前门户端唯一允许写入的字段。
    """
    if str(id) not in auth.project_ids:
        raise HTTPException(status_code=403, detail="无权访问该项目")
    project = crud.get_project(db, id)
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")

    new_val = (payload.post_delivery_person or '').strip()
    old_val = project.post_delivery_person or ''
    if new_val == old_val:
        return project

    project.post_delivery_person = new_val or None
    db.commit()
    db.refresh(project)
    crud.log_change(
        db, id,
        action_type='portal_edit',
        field_name='post_delivery_person',
        old_value=old_val, new_value=new_val,
        summary=f"门户修改交付后负责人：{_fmt_val(old_val)} → {_fmt_val(new_val)}",
    )
    return project


# ---- 变更记录（干预动作）查询接口 ----

@app.get("/api/projects/{id}/change-logs", response_model=List[schemas.ProjectChangeLogResponse])
def get_project_change_logs(id: uuid.UUID, limit: int = 200, db: Session = Depends(get_db)):
    """获取某个项目的变更记录（时间轴，倒序）"""
    project = crud.get_project(db, id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return crud.get_project_change_logs(db, id, limit=limit)


@app.get("/api/change-logs", response_model=List[schemas.GlobalChangeLogResponse])
def get_global_change_logs(limit: int = 100, db: Session = Depends(get_db)):
    """全站变更记录流（调度台右侧"干预动作"用）"""
    rows = crud.get_all_change_logs(db, limit=limit)
    result = []
    for rec, project_title in rows:
        result.append({
            "id": rec.id,
            "project_id": rec.project_id,
            "created_at": rec.created_at,
            "action_type": rec.action_type,
            "field_name": rec.field_name,
            "old_value": rec.old_value,
            "new_value": rec.new_value,
            "summary": rec.summary,
            "details": rec.details,
            "project_title": project_title,
        })
    return result


# ---- 标签分享链接功能 ----

class TagShareTokenPayload:
    """标签分享 Token 数据结构"""
    def __init__(self, tag: str, exp: datetime, iat: datetime):
        self.tag = tag
        self.exp = exp
        self.iat = iat

def create_tag_share_token(tag: str, expires_days: int = 7) -> str:
    """
    创建标签分享链接的 JWT Token
    """
    now = datetime.utcnow()
    payload = {
        "sub": "tag_share",
        "tag": tag,
        "iat": now,
        "exp": now + timedelta(days=expires_days),
        "type": "tag_share"
    }
    return jwt.encode(payload, JWT_SECRET, algorithm="HS256")

def verify_tag_share_token(token: str) -> TagShareTokenPayload:
    """
    验证标签分享 Token
    """
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
        
        if payload.get("type") != "tag_share":
            raise HTTPException(status_code=401, detail="无效的 Token 类型")
        
        return TagShareTokenPayload(
            tag=payload.get("tag", ""),
            exp=datetime.fromtimestamp(payload["exp"]),
            iat=datetime.fromtimestamp(payload["iat"])
        )
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="分享链接已过期")
    except jwt.JWTError:
        raise HTTPException(status_code=401, detail="无效的分享链接")

@app.post("/api/tags/share")
def generate_tag_share_link(payload: dict, db: Session = Depends(get_db)):
    """
    生成标签分享链接
    请求体: { "tag": "#富钢", "expires_days": 7 }
    """
    tag = payload.get("tag", "").strip()
    expires_days = payload.get("expires_days", 7)
    
    if not tag:
        raise HTTPException(status_code=400, detail="标签不能为空")
    
    # 确保标签以 # 开头
    if not tag.startswith("#"):
        tag = "#" + tag
    
    # 验证标签是否存在（至少有一个项目使用此标签）
    all_projects = crud.search_projects(db, keyword="", tags=[tag])
    if not all_projects:
        raise HTTPException(status_code=404, detail="没有找到使用该标签的项目")
    
    # 生成 Token
    token = create_tag_share_token(tag, expires_days)
    
    # 构建分享链接
    share_url = f"/share/{token}"
    
    return {
        "token": token,
        "share_url": share_url,
        "tag": tag,
        "project_count": len(all_projects),
        "expires_in": expires_days * 24 * 3600
    }

@app.get("/api/share/{token}/projects", response_model=List[schemas.ProjectResponse])
def get_shared_tag_projects(token: str, db: Session = Depends(get_db)):
    """
    通过分享链接获取标签下的项目列表（免登录）
    """
    auth = verify_tag_share_token(token)
    
    # 搜索该标签下的所有项目
    projects = crud.search_projects(db, keyword="", tags=[auth.tag])
    
    return projects

@app.get("/api/share/{token}/tag")
def get_shared_tag_info(token: str, db: Session = Depends(get_db)):
    """
    获取分享链接的标签信息（免登录）
    """
    auth = verify_tag_share_token(token)
    
    # 统计项目数量
    projects = crud.search_projects(db, keyword="", tags=[auth.tag])
    
    return {
        "tag": auth.tag,
        "project_count": len(projects),
        "expires_at": auth.exp.isoformat()
    }
