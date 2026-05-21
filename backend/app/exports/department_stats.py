"""
部门统计 Excel 导出：概览表 + 各分组子表 + 状态饼图。
"""
import io
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from ..constants import DEPARTMENT_GROUP_KEYS, STATUS_KEYS


def _write_status_block(ws, start_row: int, by_status: dict) -> int:
    """写入状态分布表，返回下一可用行号。"""
    ws.cell(row=start_row, column=1, value="状态").font = Font(bold=True)
    ws.cell(row=start_row, column=2, value="数量").font = Font(bold=True)
    r = start_row + 1
    chart_start = r
    for st in STATUS_KEYS:
        n = by_status.get(st, 0)
        if n > 0:
            ws.cell(row=r, column=1, value=st)
            ws.cell(row=r, column=2, value=n)
            r += 1
    return chart_start, r - 1


def _add_pie_chart(ws, title: str, data_start: int, data_end: int) -> None:
    if data_end < data_start:
        return
    chart = PieChart()
    chart.title = title
    labels = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
    data = Reference(ws, min_col=2, min_row=data_start, max_row=data_end)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)
    chart.width = 12
    chart.height = 8
    ws.add_chart(chart, f"D{data_start}")


def build_department_stats_xlsx(stats: Dict[str, Any], projects_by_group: Dict[str, List[dict]]) -> bytes:
    wb = Workbook()
    # --- 概览 ---
    ov = wb.active
    ov.title = "概览"
    ov["A1"] = "部门统计导出"
    ov["A1"].font = Font(bold=True, size=14)
    ov["A2"] = f"口径: {stats.get('dimension', '')}"
    ov["A3"] = f"时间: {stats.get('range', '')}"
    headers = ["分组", "合计", "实施中", "已完成", "已结项", "暂停中"]
    for c, h in enumerate(headers, 1):
        cell = ov.cell(row=5, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8F5E9")
    row = 6
    overall = stats.get("overall", {})
    ov.cell(row=row, column=1, value="整体项目")
    bs = overall.get("byStatus", {})
    ov.cell(row=row, column=2, value=overall.get("total", 0))
    for i, st in enumerate(STATUS_KEYS, 3):
        ov.cell(row=row, column=i, value=bs.get(st, 0))
    row += 1
    for g in stats.get("groups", []):
        ov.cell(row=row, column=1, value=g.get("key", ""))
        gbs = g.get("byStatus", {})
        ov.cell(row=row, column=2, value=g.get("total", 0))
        for i, st in enumerate(STATUS_KEYS, 3):
            ov.cell(row=row, column=i, value=gbs.get(st, 0))
        row += 1

    # --- 各分组子表 ---
    for group_key in DEPARTMENT_GROUP_KEYS:
        safe_name = group_key.replace("（", "(").replace("）", ")")[:31]
        ws = wb.create_sheet(title=safe_name)
        gdata = next((x for x in stats.get("groups", []) if x.get("key") == group_key), None)
        if not gdata:
            continue
        ws["A1"] = group_key
        ws["A1"].font = Font(bold=True, size=12)
        ws["A2"] = f"项目数: {gdata.get('total', 0)}"
        chart_start, chart_end = _write_status_block(ws, 4, gdata.get("byStatus", {}))
        if chart_end >= chart_start:
            _add_pie_chart(ws, f"{group_key} 状态", chart_start, chart_end)
        # 项目列表
        list_row = max(chart_end + 3, 12)
        ws.cell(row=list_row, column=1, value="项目列表").font = Font(bold=True)
        list_row += 1
        cols = ["编号", "名称", "状态", "负责人", "申报单位"]
        for c, h in enumerate(cols, 1):
            ws.cell(row=list_row, column=c, value=h).font = Font(bold=True)
        list_row += 1
        for p in projects_by_group.get(group_key, []):
            ws.cell(row=list_row, column=1, value=p.get("project_code") or "")
            ws.cell(row=list_row, column=2, value=p.get("title") or "")
            ws.cell(row=list_row, column=3, value=p.get("status") or "")
            ws.cell(row=list_row, column=4, value=p.get("leader") or "")
            ws.cell(row=list_row, column=5, value=p.get("department") or "")
            list_row += 1
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
